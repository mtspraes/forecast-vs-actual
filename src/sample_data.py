"""Synthetic source-file generator.

Writes three Excel files that reproduce the *shape and quirks* of the real
inputs, so the ingestion code has something realistic to clean:

  data/forecast_general.xlsx   early months, one sheet per month, values offset
                               into specific columns under a couple of title rows
  data/forecast_<MM>_<Month>.xlsx
                               later months as separate files, header on row 5,
                               a dynamically named month column, road lanes mixed in
  data/loading_log.xlsx        per-vehicle loading log with hand-typed market
                               names (aliases, typos, hub members by their own name)

Everything is fictional and reproducible (fixed RNG seed).
"""

from pathlib import Path
import numpy as np
from openpyxl import Workbook

from . import config

SEED = 42

# Markets that carry a forecast: maritime + hub members + road (road is later
# excluded by the transform, on purpose, to exercise that rule).
FORECAST_MARKETS = (
    list(config.MARITIME_MARKETS)
    + sorted(config.HUB_MEMBERS)
    + sorted(config.ROAD_MARKETS)
)

# Markets with a built-in systematic bias, to make the analysis findings concrete.
OVER_SHIPPERS = {"AVL", "TRN"}   # consistently ship more than forecast
UNDER_SHIPPERS = {"JZR", "ELM"}  # consistently ship less


def _target_containers(code):
    if code in config.BID_TARGETS:
        return config.BID_TARGETS[code]
    if code in config.HUB_MEMBERS:
        return 5
    if code in config.ROAD_MARKETS:
        return 8
    return 6


def generate_values(rng):
    """Return forecast (pallets) and real (vehicles) per (month, market code)."""
    forecast = {}
    real = {}
    for month, _, _ in config.MONTHS:
        for code in FORECAST_MARKETS:
            base = _target_containers(code)
            fc_containers = max(1, base * rng.uniform(0.7, 1.4))
            forecast[(month, code)] = int(round(fc_containers * config.PALLETS_PER_CONTAINER))

            # Real vehicles deviate from the forecast, with optional bias.
            factor = rng.uniform(0.8, 1.2)
            if code in OVER_SHIPPERS:
                factor *= rng.uniform(1.15, 1.4)
            elif code in UNDER_SHIPPERS:
                factor *= rng.uniform(0.55, 0.8)
            real[(month, code)] = max(0, int(round(fc_containers * factor)))
    return forecast, real


def _write_forecast_general(path, forecast, months):
    """Early months: one sheet each, market name in col E, pallets in col N,
    data starting on row 3 under two title rows."""
    wb = Workbook()
    wb.remove(wb.active)
    for month in months:
        ws = wb.create_sheet(month)
        ws["A1"] = "Export Forecast - confidential"
        ws["A2"] = f"Month: {month}"
        ws["E2"] = "Market"
        ws["N2"] = "Pallets"
        row = 3
        for code in FORECAST_MARKETS:
            ws.cell(row=row, column=5, value=code)   # col E
            ws.cell(row=row, column=14, value=forecast[(month, code)])  # col N
            row += 1
    wb.save(path)


def _write_forecast_monthly(folder, forecast, months):
    """Later months: separate files, sheet 'Forecast', header on row 5, a
    dynamically named '<MM>/2026 (Pal)' column, plus filler columns."""
    paths = []
    for month in months:
        _, year, num = next(m for m in config.MONTHS if m[0] == month)
        col_name = f"{num:02d}/{year} (Pal)"
        wb = Workbook()
        ws = wb.active
        ws.title = "Forecast"
        # Header on row 5 with some filler columns around the ones we need.
        headers = ["Line", "Region", "Channel", "Customer", "Notes", "Pais",
                   "Prev (Pal)", col_name, "Status"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h)
        pais_col = headers.index("Pais") + 1
        pal_col = headers.index(col_name) + 1
        row = 6
        for code in FORECAST_MARKETS:
            ws.cell(row=row, column=1, value=row - 5)
            ws.cell(row=row, column=pais_col, value=code)
            ws.cell(row=row, column=pal_col, value=forecast[(month, code)])
            row += 1
        path = folder / f"{num:02d}_forecast_{month.lower()}.xlsx"
        wb.save(path)
        paths.append(path)
    return paths


def _write_loading_log(path, real, rng):
    """Per-vehicle log: date, hand-typed market name, vehicle count. Spreads
    each market's monthly total across random days, using messy name variants."""
    # Reverse lookup code -> a pool of name spellings to sample from.
    spellings = {}
    for name, code in config.NAME_TO_CODE.items():
        spellings.setdefault(code, []).append(name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Loading"
    ws.append(["Date", "Market", "Vehicles"])

    rows = []
    for (month, code), total in real.items():
        if total <= 0:
            continue
        _, year, num = next(m for m in config.MONTHS if m[0] == month)
        remaining = total
        names = spellings.get(code, [code.lower()])
        while remaining > 0:
            chunk = int(min(remaining, rng.integers(1, 4)))
            day = int(rng.integers(1, 28))
            name = names[int(rng.integers(0, len(names)))]
            # Present the name with original-ish casing.
            display = name.title() if name.islower() else name
            rows.append((f"{year}-{num:02d}-{day:02d}", display, chunk))
            remaining -= chunk

    # Shuffle so the log is not grouped by market (like a real daily log).
    order = rng.permutation(len(rows))
    for i in order:
        ws.append(list(rows[i]))
    wb.save(path)


def generate_all(data_dir="data", seed=SEED):
    data_dir = Path(data_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    rng = np.random.default_rng(seed)

    forecast, real = generate_values(rng)

    early = config.MONTH_ORDER[:2]
    later = config.MONTH_ORDER[2:]

    _write_forecast_general(data_dir / "forecast_general.xlsx", forecast, early)
    _write_forecast_monthly(data_dir, forecast, later)
    _write_loading_log(data_dir / "loading_log.xlsx", real, rng)

    return data_dir


if __name__ == "__main__":
    out = generate_all()
    print(f"Synthetic source files written to: {out.resolve()}")
