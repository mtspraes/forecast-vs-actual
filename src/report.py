"""Reporting - the operator deliverables: a formatted Excel workbook and charts.

The Excel mirrors the real tool's output: one "Comparison <Month>" sheet each
(Market, Forecast, Actual, Difference, Target) with a combo chart (forecast vs.
actual bars plus a target line), and a summary sheet of the KPIs. Matplotlib
versions of the same charts are saved as PNGs for the README and notebook.
"""

from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # headless rendering
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker

from . import config, analysis

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT = Font(bold=True, size=11)
_thin = Side(style="thin")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _month_rows(comp, month):
    m = comp[comp["month"] == month].sort_values("market")
    return list(m.itertuples(index=False))


def write_excel(comp, path="output/forecast_vs_actual.xlsx"):
    """Write the formatted workbook. Returns the output path."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    _write_summary_sheet(wb, comp)

    months = [m for m in config.MONTH_ORDER if (comp["month"] == m).any()]
    for month in months:
        _write_comparison_sheet(wb, comp, month)

    wb.save(path)
    return path


def _write_summary_sheet(wb, comp):
    ws = wb.create_sheet("Summary")
    monthly = analysis.monthly_kpis(comp)
    overall = analysis.overall_kpis(comp)

    ws["A1"] = "Forecast vs. Actual - KPI Summary"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Month", "Forecast (cnt)", "Actual (veh)", "Bias", "Bias %", "MAPE %", "Hit rate %"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    r = 4
    for row in monthly.itertuples(index=False):
        ws.cell(r, 1, row.month)
        ws.cell(r, 2, row.forecast_containers)
        ws.cell(r, 3, row.real_vehicles)
        ws.cell(r, 4, row.bias)
        ws.cell(r, 5, row.bias_pct)
        ws.cell(r, 6, row.mape_pct)
        ws.cell(r, 7, row.hit_rate_pct)
        r += 1
    # Overall line.
    ws.cell(r, 1, "OVERALL").font = TOTAL_FONT
    for c, key in zip(range(2, 8), ["forecast_containers", "real_vehicles", "bias", "bias_pct", "mape_pct", "hit_rate_pct"]):
        cell = ws.cell(r, c, overall[key])
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    ws.column_dimensions["A"].width = 16


def _write_comparison_sheet(wb, comp, month):
    ws = wb.create_sheet(f"Comparison {month}")
    headers = ["Market", "Forecast (cnt)", "Actual (veh)", "Difference", "Target"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if c <= 4:
            cell.border = BORDER

    rows = _month_rows(comp, month)
    first = 2
    r = first
    for row in rows:
        ws.cell(r, 1, row.market).border = BORDER
        cB = ws.cell(r, 2, int(round(row.forecast_containers))); cB.border = BORDER; cB.number_format = "0"
        cC = ws.cell(r, 3, int(round(row.real_vehicles))); cC.border = BORDER; cC.number_format = "0"
        cD = ws.cell(r, 4, f"=C{r}-B{r}"); cD.border = BORDER; cD.number_format = "0"
        ws.cell(r, 5, int(row.bid))  # target column, no border (matches model)
        r += 1
    last = r - 1

    # TOTAL row (one blank row above).
    r += 1
    ws.cell(r, 1, "TOTAL").font = TOTAL_FONT
    ws.cell(r, 1).fill = TOTAL_FILL
    for col, formula in [(2, f"=SUM(B{first}:B{last})"), (3, f"=SUM(C{first}:C{last})"), (4, f"=C{r}-B{r}")]:
        cell = ws.cell(r, col, formula)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = "0"
    ws.column_dimensions["A"].width = 18

    _add_combo_chart(ws, month, first, last)


def _add_combo_chart(ws, month, first, last):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"Comparison {month}"
    chart.width = 15
    chart.height = 7.5
    chart.legend.position = "r"
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(cats)

    line = LineChart()
    line.add_data(Reference(ws, min_col=5, min_row=1, max_row=last), titles_from_data=True)
    line.set_categories(cats)
    line.series[0].marker = Marker(symbol="none")
    chart += line
    ws.add_chart(chart, "G1")


# --- Matplotlib charts (for README / notebook) ---

def plot_month_combo(comp, month, ax=None):
    rows = _month_rows(comp, month)
    markets = [r.market for r in rows]
    forecast = [r.forecast_containers for r in rows]
    actual = [r.real_vehicles for r in rows]
    bid = [r.bid for r in rows]

    own_ax = ax is None
    if own_ax:
        fig, ax = plt.subplots(figsize=(10, 5))
    x = range(len(markets))
    w = 0.4
    ax.bar([i - w / 2 for i in x], forecast, width=w, label="Forecast (cnt)", color="#2F5496")
    ax.bar([i + w / 2 for i in x], actual, width=w, label="Actual (veh)", color="#9DC3E6")
    ax.plot(list(x), bid, color="#7030A0", marker="o", linewidth=2, label="Target (BID)")
    ax.set_xticks(list(x))
    ax.set_xticklabels(markets, rotation=45, ha="right")
    ax.set_ylabel("Containers / vehicles")
    ax.set_title(f"Forecast vs. Actual - {month}")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    if own_ax:
        fig.tight_layout()
        return fig
    return ax


def plot_accuracy_trend(comp, ax=None):
    monthly = analysis.monthly_kpis(comp)
    own_ax = ax is None
    if own_ax:
        fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(monthly["month"], monthly["hit_rate_pct"], marker="o", color="#2F5496", label="Hit rate %")
    ax.plot(monthly["month"], monthly["mape_pct"], marker="s", color="#C00000", label="MAPE %")
    ax.axhline(0, color="gray", linewidth=0.8)
    ax.plot(monthly["month"], monthly["bias_pct"], marker="^", color="#7030A0", label="Bias %")
    ax.set_ylabel("%")
    ax.set_title("Forecast accuracy over time")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    if own_ax:
        fig.tight_layout()
        return fig
    return ax


def save_charts(comp, out_dir="output/charts"):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for month in [m for m in config.MONTH_ORDER if (comp["month"] == m).any()]:
        fig = plot_month_combo(comp, month)
        p = out_dir / f"comparison_{month.lower()}.png"
        fig.savefig(p, dpi=110)
        plt.close(fig)
        saved.append(p)
    fig = plot_accuracy_trend(comp)
    p = out_dir / "accuracy_trend.png"
    fig.savefig(p, dpi=110)
    plt.close(fig)
    saved.append(p)
    return saved
